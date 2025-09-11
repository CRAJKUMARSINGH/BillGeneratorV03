"""
Enhanced utilities module combining best features from all versions
Includes comprehensive validation, formatting, and helper functions
"""

import pandas as pd
import openpyxl
from typing import Any, Union, Dict, List, Optional
from datetime import datetime
import logging
import re
import os
from pathlib import Path

logger = logging.getLogger(__name__)

def validate_excel_file(uploaded_file) -> Dict[str, Any]:
    """
    Enhanced Excel file validation with comprehensive checks
    """
    validation_result = {
        'valid': True,
        'error': None,
        'warnings': [],
        'file_info': {}
    }
    
    try:
        # Reset file pointer
        uploaded_file.seek(0)
        
        # Basic file checks
        file_size = len(uploaded_file.getvalue())
        validation_result['file_info']['size_mb'] = file_size / (1024 * 1024)
        validation_result['file_info']['filename'] = uploaded_file.name
        
        # Check file size limits (50MB max for performance)
        if file_size > 50 * 1024 * 1024:
            validation_result['valid'] = False
            validation_result['error'] = f"File too large ({file_size / (1024*1024):.1f}MB). Maximum 50MB allowed."
            return validation_result
        
        # Try to load the workbook
        try:
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        except Exception as e:
            validation_result['valid'] = False
            validation_result['error'] = f"Invalid Excel file format: {str(e)}"
            return validation_result
        
        # Check if file has any sheets
        if not workbook.sheetnames:
            validation_result['valid'] = False
            validation_result['error'] = "Excel file contains no sheets"
            return validation_result
        
        validation_result['file_info']['sheets'] = workbook.sheetnames
        validation_result['file_info']['sheet_count'] = len(workbook.sheetnames)
        
        # Check for minimum required sheets with flexible matching
        sheet_names_lower = [name.lower() for name in workbook.sheetnames]
        
        required_sheets = {
            'title': ['title', 'cover', 'front', 'project', 'header'],
            'work_order': ['work order', 'work_order', 'workorder', 'wo', 'order'],
            'bill_quantity': ['bill quantity', 'bill_quantity', 'billquantity', 'bq', 'quantity', 'bill']
        }
        
        found_sheets = {}
        missing_sheets = []
        
        for category, keywords in required_sheets.items():
            found = False
            for keyword in keywords:
                for sheet_name in sheet_names_lower:
                    if keyword in sheet_name or sheet_name in keyword:
                        found_sheets[category] = workbook.sheetnames[sheet_names_lower.index(sheet_name)]
                        found = True
                        break
                if found:
                    break
            
            if not found:
                missing_sheets.append(category.replace('_', ' ').title())
        
        validation_result['file_info']['found_sheets'] = found_sheets
        
        if missing_sheets:
            validation_result['warnings'].append(f"Potentially missing sheets: {', '.join(missing_sheets)}")
        
        # Test reading at least one sheet for data validation
        try:
            first_sheet = workbook.sheetnames[0]
            df_test = pd.read_excel(uploaded_file, sheet_name=first_sheet, nrows=10)
            
            if df_test.empty:
                validation_result['warnings'].append(f"Sheet '{first_sheet}' appears to be empty")
            
            validation_result['file_info']['columns_sample'] = df_test.columns.tolist()[:10]
            
        except Exception as e:
            validation_result['warnings'].append(f"Could not read sheet data: {str(e)}")
        
        # Additional validations
        if validation_result['file_info']['sheet_count'] < 3:
            validation_result['warnings'].append("File has fewer sheets than typically expected (minimum 3)")
        
        # Check for common issues
        for sheet_name in workbook.sheetnames:
            if 'temp' in sheet_name.lower() or 'backup' in sheet_name.lower():
                validation_result['warnings'].append(f"Found temporary/backup sheet: {sheet_name}")
        
        logger.info(f"Excel file validation successful. File: {uploaded_file.name}, Sheets: {len(workbook.sheetnames)}")
        
    except Exception as e:
        validation_result['valid'] = False
        validation_result['error'] = f"Validation error: {str(e)}"
        logger.error(f"File validation error: {str(e)}")
    
    finally:
        # Reset file pointer for future use
        uploaded_file.seek(0)
    
    return validation_result

def safe_float_conversion(value: Any, default: float = 0.0) -> float:
    """
    Enhanced safe float conversion with better error handling
    """
    if value is None or pd.isna(value):
        return default
    
    # Handle numeric types
    if isinstance(value, (int, float)):
        # Check for infinity and NaN
        import math
        if pd.isna(value) or not math.isfinite(value):
            return default
        return float(value)
    
    # Handle string values
    if isinstance(value, str):
        value_clean = value.strip().lower()
        
        # Handle specific text cases
        text_to_zero = ['above', 'n/a', 'na', 'nil', '', '-', 'null', 'none', 'tbc', 'tbd']
        if value_clean in text_to_zero:
            return default
        
        # Handle currency symbols and formatting
        value_clean = value.strip()
        
        # Remove common currency symbols and formatting
        currency_chars = ['₹', '$', '€', '£', '¥', 'Rs', 'Rs.', 'INR']
        for char in currency_chars:
            value_clean = value_clean.replace(char, '')
        
        # Remove thousands separators
        value_clean = value_clean.replace(',', '').replace(' ', '')
        
        # Handle percentage
        is_percentage = value_clean.endswith('%')
        if is_percentage:
            value_clean = value_clean.rstrip('%')
        
        # Extract numeric value using regex
        try:
            # Match decimal numbers (including negative)
            numeric_pattern = r'[-+]?(?:\d+\.?\d*|\.\d+)(?:[eE][-+]?\d+)?'
            matches = re.findall(numeric_pattern, value_clean)
            
            if matches:
                numeric_value = float(matches[0])
                
                # Apply percentage conversion
                if is_percentage:
                    numeric_value = numeric_value / 100
                
                return numeric_value
            else:
                logger.warning(f"Could not extract numeric value from '{value}', returning {default}")
                return default
                
        except (ValueError, TypeError, OverflowError):
            logger.warning(f"Could not convert '{value}' to float, returning {default}")
            return default
    
    # Try direct conversion for other types
    try:
        result = float(value)
        import math
        if pd.isna(result) or not math.isfinite(result):
            return default
        return result
    except (ValueError, TypeError, OverflowError):
        logger.warning(f"Could not convert '{value}' of type {type(value)} to float, returning {default}")
        return default

def round_to_nearest(value: Union[int, float], decimal_places: int = 0, rounding_rule: str = 'standard') -> float:
    """
    Enhanced rounding with different rounding rules for government compliance
    
    Args:
        value: Number to round
        decimal_places: Number of decimal places
        rounding_rule: 'standard', 'up', 'down', 'even' (banker's rounding)
    """
    if not isinstance(value, (int, float)) or pd.isna(value):
        return 0.0
    
    try:
        if rounding_rule == 'up':
            import math
            factor = 10 ** decimal_places
            return math.ceil(value * factor) / factor
        elif rounding_rule == 'down':
            import math
            factor = 10 ** decimal_places
            return math.floor(value * factor) / factor
        elif rounding_rule == 'even':
            # Banker's rounding - round to nearest even
            from decimal import Decimal, ROUND_HALF_EVEN
            decimal_value = Decimal(str(value))
            return float(decimal_value.quantize(Decimal('0.1') ** decimal_places, rounding=ROUND_HALF_EVEN))
        else:
            # Standard rounding
            return round(float(value), decimal_places)
            
    except (ValueError, TypeError, OverflowError):
        logger.warning(f"Error rounding value {value}, returning 0.0")
        return 0.0

def format_currency(amount: Union[int, float], currency: str = '₹', include_decimals: bool = True) -> str:
    """
    Enhanced currency formatting with Indian numbering system support
    """
    import math
    if pd.isna(amount) or not math.isfinite(amount):
        return f"{currency}0{'.00' if include_decimals else ''}"
    
    try:
        abs_amount = abs(float(amount))
        
        # Indian numbering system formatting
        if abs_amount >= 10000000:  # 1 crore
            if include_decimals:
                formatted = f"{abs_amount:,.2f}"
            else:
                formatted = f"{abs_amount:,.0f}"
        elif abs_amount >= 100000:  # 1 lakh
            if include_decimals:
                formatted = f"{abs_amount:,.2f}"
            else:
                formatted = f"{abs_amount:,.0f}"
        else:
            if include_decimals:
                formatted = f"{abs_amount:,.2f}"
            else:
                formatted = f"{abs_amount:,.0f}"
        
        # Add negative sign if needed
        if amount < 0:
            formatted = f"-{currency}{formatted}"
        else:
            formatted = f"{currency}{formatted}"
        
        return formatted
        
    except (ValueError, TypeError):
        return f"{currency}0{'.00' if include_decimals else ''}"

def format_date(date_value: Any, format_string: str = "%d/%m/%Y", default: str = "") -> str:
    """
    Enhanced date formatting with multiple input format support
    """
    # Support legacy calling style where the second argument is actually the default value
    # Example: format_date("invalid", "default")
    if isinstance(format_string, str) and '%' not in format_string and default == "":
        default = format_string
        format_string = "%d/%m/%Y"

    if date_value is None or pd.isna(date_value):
        return default
    
    # Handle datetime objects
    if isinstance(date_value, (datetime, pd.Timestamp)):
        try:
            return date_value.strftime(format_string)
        except (ValueError, AttributeError):
            return default
    
    # Handle string dates
    if isinstance(date_value, str):
        date_str = date_value.strip()
        if not date_str:
            return default
        
        # Common date formats to try
        date_formats = [
            "%Y-%m-%d",           # 2023-12-25
            "%d/%m/%Y",           # 25/12/2023
            "%d-%m-%Y",           # 25-12-2023
            "%m/%d/%Y",           # 12/25/2023
            "%Y/%m/%d",           # 2023/12/25
            "%d %B %Y",           # 25 December 2023
            "%B %d, %Y",          # December 25, 2023
            "%d-%b-%Y",           # 25-Dec-2023
        ]
        
        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(date_str, fmt)
                return parsed_date.strftime(format_string)
            except ValueError:
                continue

        # If no format worked, return the provided default
        return default if default is not None else ""
    
    # Handle numeric dates (Excel serial numbers)
    if isinstance(date_value, (int, float)):
        try:
            # Excel date serial number conversion
            if 1 <= date_value <= 2958465:  # Valid Excel date range
                base_date = datetime(1899, 12, 30)  # Excel's base date
                parsed_date = base_date + pd.Timedelta(days=date_value)
                return parsed_date.strftime(format_string)
        except (ValueError, OverflowError):
            pass

    return default

def clean_text(text: Any, max_length: int = None) -> str:
    """
    Enhanced text cleaning and standardization
    """
    if text is None or pd.isna(text):
        return ""
    
    # Convert to string
    clean_text = str(text).strip()
    
    # Remove extra whitespace
    clean_text = re.sub(r'\s+', ' ', clean_text)
    
    # Remove control characters
    clean_text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', clean_text)
    
    # Remove special quotes and replace with standard ones
    clean_text = clean_text.replace('"', '"').replace('"', '"')
    clean_text = clean_text.replace(''', "'").replace(''', "'")
    
    # Limit length if specified
    if max_length and len(clean_text) > max_length:
        clean_text = clean_text[:max_length-3] + "..."
    
    return clean_text

def validate_numeric_value(value: Any, min_value: float = None, max_value: float = None) -> Dict[str, Any]:
    """
    Enhanced numeric validation with detailed feedback
    """
    result = {
        'valid': False,
        'value': None,
        'original': value,
        'warnings': []
    }
    
    try:
        numeric_value = safe_float_conversion(value)
        result['value'] = numeric_value
        
        # Check range constraints
        if min_value is not None and numeric_value < min_value:
            result['warnings'].append(f"Value {numeric_value} is below minimum {min_value}")
            return result
        
        if max_value is not None and numeric_value > max_value:
            result['warnings'].append(f"Value {numeric_value} exceeds maximum {max_value}")
            return result
        
        # Check for reasonable business values
        if numeric_value < 0:
            result['warnings'].append("Negative value detected")
        
        if numeric_value > 1000000000:  # 100 crores
            result['warnings'].append("Unusually large value detected")
        
        result['valid'] = True
        
    except Exception as e:
        result['warnings'].append(f"Validation error: {str(e)}")
    
    return result

def extract_project_info(title_data: Dict[str, Any]) -> Dict[str, str]:
    """
    Enhanced project information extraction and standardization
    """
    project_info = {}
    
    # Field mappings with cleaning
    field_mappings = {
        'project_name': clean_text(title_data.get('project_name', '')),
        'contractor_name': clean_text(title_data.get('contractor_name', '')),
        'agreement_no': clean_text(title_data.get('agreement_no', '')),
        'work_order_no': clean_text(title_data.get('work_order_no', '')),
        'location': clean_text(title_data.get('location', '')),
        'division': clean_text(title_data.get('division', '')),
        'estimated_cost': safe_float_conversion(title_data.get('estimated_cost', 0)),
        'start_date': format_date(title_data.get('start_date', '')),
        'completion_date': format_date(title_data.get('completion_date', ''))
    }
    
    # Only include non-empty values
    for key, value in field_mappings.items():
        if value:  # Non-empty string or non-zero number
            project_info[key] = value
    
    return project_info

def calculate_gst(amount: float, gst_rate: float = 18.0) -> Dict[str, float]:
    """
    Enhanced GST calculation with proper rounding
    """
    try:
        base_amount = safe_float_conversion(amount)
        rate = safe_float_conversion(gst_rate)
        
        if base_amount <= 0 or rate < 0:
            return {
                'base_amount': 0.0,
                'gst_rate': rate,
                'gst_amount': 0.0,
                'total_with_gst': 0.0
            }
        
        # Calculate GST with proper rounding
        gst_amount = round_to_nearest(base_amount * (rate / 100), 2)
        total_with_gst = base_amount + gst_amount
        
        return {
            'base_amount': base_amount,
            'gst_rate': rate,
            'gst_amount': gst_amount,
            'total_with_gst': total_with_gst
        }
        
    except Exception as e:
        logger.error(f"Error calculating GST: {str(e)}")
        return {
            'base_amount': 0.0,
            'gst_rate': 18.0,
            'gst_amount': 0.0,
            'total_with_gst': 0.0
        }

def generate_serial_number(index: int, prefix: str = "", padding: int = 3) -> str:
    """
    Enhanced serial number generation
    """
    try:
        if prefix:
            return f"{prefix}{index:0{padding}d}"
        else:
            return f"{index:0{padding}d}"
    except (ValueError, TypeError):
        return str(index)

def sanitize_filename(filename: str) -> str:
    """
    Enhanced filename sanitization for cross-platform compatibility
    """
    if not filename:
        return "document"
    
    # Remove or replace invalid characters (preserve number of invalids as underscores)
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)

    # Remove control characters
    filename = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', filename)

    # Collapse only whitespace to single space, then convert spaces to underscores (do NOT collapse underscores)
    filename = re.sub(r'\s+', ' ', filename.strip())
    filename = filename.replace(' ', '_')
    
    # Remove leading/trailing dots and spaces
    filename = filename.strip('. ')
    
    # Limit length (Windows has 260 char limit for full path)
    if len(filename) > 100:
        name, ext = os.path.splitext(filename)
        filename = name[:95] + ext
    
    # Ensure filename is not empty
    if not filename:
        filename = "document"
    
    return filename

def get_timestamp(format_type: str = "file") -> str:
    """
    Enhanced timestamp generation with different formats
    """
    now = datetime.now()
    
    formats = {
        'file': "%Y%m%d_%H%M%S",           # 20231225_143022
        'display': "%d/%m/%Y %H:%M:%S",    # 25/12/2023 14:30:22
        'iso': now.isoformat(),            # 2023-12-25T14:30:22
        'compact': "%Y%m%d%H%M%S",         # 20231225143022
        'date_only': "%Y%m%d",             # 20231225
        'readable': "%d %B %Y at %I:%M %p" # 25 December 2023 at 02:30 PM
    }
    
    if format_type in formats:
        if format_type == 'iso':
            return formats[format_type]
        else:
            return now.strftime(formats[format_type])
    else:
        return now.strftime("%Y%m%d_%H%M%S")

def log_processing_info(operation: str, records_processed: int, errors: int = 0, duration: float = None):
    """
    Enhanced logging for processing operations
    """
    message = f"{operation}: {records_processed} records processed"
    
    if errors > 0:
        message += f", {errors} errors"
    
    if duration:
        message += f", completed in {duration:.2f}s"
    
    logger.info(message)

def validate_sheet_structure(df: pd.DataFrame, required_columns: List[str], sheet_name: str = "") -> Dict[str, Any]:
    """
    Enhanced DataFrame structure validation
    """
    result = {
        'valid': True,
        'missing_columns': [],
        'extra_columns': [],
        'warnings': [],
        'suggestions': [],
        'column_mapping': {}
    }
    
    if df.empty:
        result['warnings'].append(f"Sheet {sheet_name} is empty")
        result['valid'] = False
        return result
    
    # Normalize column names for comparison
    df_columns_lower = [col.lower().strip() for col in df.columns]
    required_columns_lower = [col.lower().strip() for col in required_columns]
    
    # Check for missing required columns with fuzzy matching
    for req_col in required_columns:
        req_col_lower = req_col.lower().strip()
        exact_match = req_col_lower in df_columns_lower
        
        if not exact_match:
            # Try fuzzy matching
            fuzzy_matches = []
            for df_col in df.columns:
                df_col_lower = df_col.lower().strip()
                # Check for partial matches
                if req_col_lower in df_col_lower or df_col_lower in req_col_lower:
                    fuzzy_matches.append(df_col)
                # Check for keyword matches
                req_words = req_col_lower.replace('_', ' ').split()
                df_words = df_col_lower.replace('_', ' ').split()
                if any(word in df_words for word in req_words):
                    fuzzy_matches.append(df_col)
            
            if fuzzy_matches:
                result['suggestions'].append(f"Column '{req_col}' not found, possible matches: {fuzzy_matches}")
                result['column_mapping'][req_col] = fuzzy_matches[0]  # Use first match
            else:
                result['missing_columns'].append(req_col)
                result['valid'] = False
        else:
            # Exact match found
            original_col = df.columns[df_columns_lower.index(req_col_lower)]
            result['column_mapping'][req_col] = original_col
    
    # Identify extra columns
    for df_col in df.columns:
        df_col_lower = df_col.lower().strip()
        if df_col_lower not in required_columns_lower:
            # Check if it's a close match to required columns
            is_close_match = False
            for req_col_lower in required_columns_lower:
                if df_col_lower in req_col_lower or req_col_lower in df_col_lower:
                    is_close_match = True
                    break
            
            if not is_close_match:
                result['extra_columns'].append(df_col)
    
    # Additional data quality checks
    if len(df) < 2:
        result['warnings'].append("Sheet has very few rows of data")
    
    # Check for empty columns
    empty_columns = [col for col in df.columns if df[col].isna().all()]
    if empty_columns:
        result['warnings'].append(f"Empty columns found: {empty_columns}")
    
    return result

def create_backup_filename(original_path: str) -> str:
    """
    Create backup filename with timestamp
    """
    path = Path(original_path)
    timestamp = get_timestamp('compact')
    return str(path.parent / f"{path.stem}_backup_{timestamp}{path.suffix}")

def ensure_directory_exists(directory_path: str) -> bool:
    """
    Ensure directory exists, create if it doesn't
    """
    try:
        Path(directory_path).mkdir(parents=True, exist_ok=True)
        return True
    except Exception as e:
        logger.error(f"Error creating directory {directory_path}: {str(e)}")
        return False
