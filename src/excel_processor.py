import pandas as pd
import openpyxl
from typing import Dict, List, Any, Optional, Tuple
import logging
from datetime import datetime, date
import numpy as np
try:
    from src.utils import safe_float_conversion, format_currency, round_to_nearest, clean_text
except ImportError:
    try:
        from utils import safe_float_conversion, format_currency, round_to_nearest, clean_text
    except ImportError:
        # Fallback functions if utils module is not available
        def safe_float_conversion(value, default=0.0):
            try:
                return float(value) if value is not None else default
            except (ValueError, TypeError):
                return default
        
        def format_currency(amount):
            return f"₹{amount:,.2f}"
        
        def round_to_nearest(value, nearest=0.01):
            return round(value / nearest) * nearest
        
        def clean_text(text):
            return str(text).strip() if text is not None else ""

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """
    Enhanced Excel processor combining best features from all versions.
    Handles both old and new pattern files with comprehensive validation and error handling.
    """
    
    def __init__(self, excel_file=None):
        """Initialize with uploaded Excel file"""
        self.excel_file = excel_file
        self.workbook = None
        self.processed_data = {}
        self.file_pattern = None  # 'old' or 'new'
        self.validation_warnings = []
        self.data = {}  # For compatibility
        
    def load_workbook(self) -> bool:
        """Load Excel workbook and detect file pattern"""
        try:
            self.workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
            self.detect_file_pattern()
            logger.info(f"Workbook loaded successfully. Pattern: {self.file_pattern}")
            return True
        except Exception as e:
            logger.error(f"Error loading workbook: {str(e)}")
            raise Exception(f"Failed to load Excel file: {str(e)}")
    
    def detect_file_pattern(self):
        """Enhanced pattern detection with multiple criteria"""
        sheet_names = [name.lower() for name in self.workbook.sheetnames]
        
        # Check for typical new pattern indicators
        new_pattern_indicators = ['title', 'cover', 'front', 'project']
        old_pattern_indicators = ['sheet1', 'data', 'main']
        
        new_score = sum(1 for indicator in new_pattern_indicators 
                       if any(indicator in name for name in sheet_names))
        old_score = sum(1 for indicator in old_pattern_indicators 
                       if any(indicator in name for name in sheet_names))
        
        # Additional checks for structure
        if len(self.workbook.sheetnames) > 3 and new_score > 0:
            self.file_pattern = 'new'
        elif old_score > new_score:
            self.file_pattern = 'old'
        else:
            # Default to new pattern for better compatibility
            self.file_pattern = 'new'
        
        logger.info(f"Detected file pattern: {self.file_pattern} (new_score: {new_score}, old_score: {old_score})")
    
    def get_required_sheets(self) -> Dict[str, List[str]]:
        """Get mapping of required sheets with flexible matching"""
        return {
            'first_page': ['first page', 'first_page', 'front', 'title', 'cover', 'front', 'project', 'header'],
            'work_order': ['work order', 'work_order', 'workorder', 'wo', 'order'],
            'bill_quantity': ['bill quantity', 'bill_quantity', 'billquantity', 'bq', 'quantity', 'bill'],
            'extra_items': ['extra items', 'extra_items', 'extraitems', 'extra', 'additional'],
            'deviation_statement': ['deviation statement', 'deviation_statement', 'deviation'],
            'note_sheet': ['note sheet', 'note_sheet', 'notes']
        }
    
    def find_sheet_by_keywords(self, keywords: List[str]) -> Optional[str]:
        """Find sheet by keyword matching with fuzzy logic"""
        available_sheets = [name.lower() for name in self.workbook.sheetnames]
        
        # Exact match first
        for keyword in keywords:
            for sheet in self.workbook.sheetnames:
                if keyword.lower() == sheet.lower():
                    return sheet
        
        # Partial match
        for keyword in keywords:
            for sheet in self.workbook.sheetnames:
                if keyword.lower() in sheet.lower():
                    return sheet
        
        # Reverse partial match
        for sheet in self.workbook.sheetnames:
            for keyword in keywords:
                if sheet.lower() in keyword.lower():
                    return sheet
        
        return None
    
    def validate_sheets(self) -> Dict[str, Any]:
        """Enhanced sheet validation with detailed feedback"""
        if not self.workbook:
            self.load_workbook()
            
        available_sheets = self.workbook.sheetnames
        required_sheets = self.get_required_sheets()
        
        validation_result = {
            'valid': True,
            'missing_sheets': [],
            'available_sheets': available_sheets,
            'pattern': self.file_pattern,
            'sheet_mapping': {},
            'warnings': []
        }
        
        # Map available sheets to required categories
        for category, keywords in required_sheets.items():
            found_sheet = self.find_sheet_by_keywords(keywords)
            if found_sheet:
                validation_result['sheet_mapping'][category] = found_sheet
            elif category != 'extra_items':  # Extra Items is optional
                validation_result['missing_sheets'].append(category)
                validation_result['valid'] = False
        
        # Additional validation warnings
        if len(available_sheets) < 3:
            validation_result['warnings'].append("File has fewer than expected sheets")
        
        return validation_result
    
    def process_all_sheets(self) -> Dict[str, Any]:
        """Process all sheets with comprehensive error handling"""
        try:
            if not self.workbook:
                self.load_workbook()
            
            validation = self.validate_sheets()
            if not validation['valid']:
                logger.error(f"Sheet validation failed: {validation['missing_sheets']}")
                return {}
            
            processed_data = {}
            
            # Process each sheet type
            sheet_mapping = validation.get('sheet_mapping', {})
            
            # Process Title sheet (First Page)
            if 'first_page' in sheet_mapping:
                processed_data['title'] = self.process_title_sheet(sheet_mapping['first_page'])
            
            # Process Work Order sheet
            if 'work_order' in sheet_mapping:
                processed_data['work_order'] = self.process_work_order_sheet(sheet_mapping['work_order'])
            
            # Process Bill Quantity sheet
            if 'bill_quantity' in sheet_mapping:
                processed_data['bill_quantity'] = self.process_bill_quantity_sheet(sheet_mapping['bill_quantity'])
            
            # Process Extra Items sheet (if exists)
            if 'extra_items' in sheet_mapping:
                processed_data['extra_items'] = self.process_extra_items_sheet(sheet_mapping['extra_items'])
            else:
                processed_data['extra_items'] = []
            
            # Calculate totals and financial summary
            processed_data['totals'] = self.calculate_financial_totals(processed_data)
            
            # Generate Note Sheet content
            processed_data['note_sheet_content'] = self.generate_note_sheet_content(processed_data)
            
            logger.info(f"Successfully processed {len(processed_data)} data categories")
            return processed_data
            
        except Exception as e:
            logger.error(f"Error processing sheets: {str(e)}")
            return {}
    
    def process_title_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """Enhanced title sheet processing with multiple patterns"""
        try:
            worksheet = self.workbook[sheet_name]
            title_data = {}
            
            # Enhanced field mapping with multiple variations
            field_mappings = {
                'project_name': ['project', 'project name', 'work name', 'name of work', 'scheme'],
                'contractor_name': ['contractor', 'contractor name', 'agency', 'firm', 'company'],
                'agreement_no': ['agreement', 'agreement no', 'agreement number', 'contract no'],
                'work_order_no': ['work order', 'work order no', 'wo no', 'order no'],
                'location': ['location', 'site', 'place', 'district'],
                'estimated_cost': ['estimated cost', 'estimate', 'cost', 'amount'],
                'start_date': ['start date', 'commencement', 'begin date'],
                'completion_date': ['completion date', 'end date', 'target date']
            }
            
            # Process rows looking for field matches
            for row in worksheet.iter_rows(max_row=30, values_only=True):
                if not row or not row[0]:
                    continue
                
                cell_value = str(row[0]).strip().lower()
                
                for field_name, keywords in field_mappings.items():
                    for keyword in keywords:
                        if keyword in cell_value and len(row) > 1 and row[1]:
                            value = row[1]
                            # Handle date fields
                            if 'date' in field_name and isinstance(value, (datetime, date)):
                                title_data[field_name] = value.strftime("%d/%m/%Y")
                            else:
                                title_data[field_name] = clean_text(str(value))
                            break
                    if field_name in title_data:
                        break
            
            logger.info(f"Processed title sheet: {len(title_data)} fields extracted")
            return title_data
            
        except Exception as e:
            logger.error(f"Error processing title sheet: {str(e)}")
            return {}
    
    def process_work_order_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Enhanced work order processing with flexible column detection"""
        try:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            
            # Clean column names
            df.columns = df.columns.str.strip().str.lower()
            
            # Enhanced column mapping
            column_mappings = {
                'serial_no': ['s.no', 'serial', 'sr.no', 'no', 'item no', 'sl.no'],
                'description': ['description', 'particulars', 'item', 'work', 'details'],
                'unit': ['unit', 'units', 'measurement', 'measure'],
                'quantity': ['quantity', 'qty', 'amount', 'nos'],
                'rate': ['rate', 'unit rate', 'price', 'cost'],
                'amount': ['amount', 'total', 'value', 'cost'],
                'remark': ['remark', 'remarks', 'note', 'comment']
            }
            
            # Map columns
            mapped_columns = {}
            for target_col, possible_names in column_mappings.items():
                for possible_name in possible_names:
                    for actual_col in df.columns:
                        if possible_name in actual_col:
                            mapped_columns[target_col] = actual_col
                            break
                    if target_col in mapped_columns:
                        break
            
            work_items = []
            for index, row in df.iterrows():
                # Skip header rows and empty rows
                if index < 5:  # Allow for headers in first few rows
                    desc_value = row.get(mapped_columns.get('description', ''), '')
                    if pd.isna(desc_value) or str(desc_value).strip() == '':
                        continue
                
                # Only process rows with meaningful data
                description = clean_text(row.get(mapped_columns.get('description', ''), ''))
                if not description or len(description) < 3:
                    continue
                
                # Get quantity and rate values
                quantity = safe_float_conversion(row.get(mapped_columns.get('quantity', ''), 0))
                rate = safe_float_conversion(row.get(mapped_columns.get('rate', ''), 0))
                
                # Handle blank/zero rate logic according to VBA specification
                # If rate is blank or zero, only populate serial_no and description
                if rate == 0 or pd.isna(row.get(mapped_columns.get('rate', ''))) or str(row.get(mapped_columns.get('rate', ''))).strip() == '':
                    item = {
                        'serial_no': clean_text(row.get(mapped_columns.get('serial_no', ''), str(index + 1))),
                        'description': description,
                        'unit': '',  # Keep blank when rate is blank/zero
                        'quantity': 0,  # Keep zero when rate is blank/zero
                        'rate': 0,  # Keep zero when rate is blank/zero
                        'amount': 0,  # Keep zero when rate is blank/zero
                        'remark': ''  # Keep blank when rate is blank/zero
                    }
                else:
                    # Normal processing when rate is not blank/zero
                    item = {
                        'serial_no': clean_text(row.get(mapped_columns.get('serial_no', ''), str(index + 1))),
                        'description': description,
                        'unit': clean_text(row.get(mapped_columns.get('unit', ''), ''),
                        'quantity': quantity,
                        'rate': rate,
                        'remark': clean_text(row.get(mapped_columns.get('remark', ''), ''))
                    }
                    
                    # Calculate amount if not provided
                    if 'amount' in mapped_columns:
                        item['amount'] = safe_float_conversion(row.get(mapped_columns['amount'], 0))
                    else:
                        item['amount'] = item['quantity'] * item['rate']
                    
                    # Apply rounding rules
                    item['amount'] = round_to_nearest(item['amount'])
                
                work_items.append(item)
            
            logger.info(f"Processed {len(work_items)} work order items")
            return work_items
            
        except Exception as e:
            logger.error(f"Error processing work order sheet: {str(e)}")
            return []

    def process_bill_quantity_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Enhanced bill quantity processing with smart data detection"""
        try:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            
            # Clean and standardize column names
            df.columns = df.columns.str.strip().str.lower()
            
            # Enhanced column mapping for bill quantity
            column_mappings = {
                'serial_no': ['s.no', 'serial', 'sr.no', 'no', 'item no', 'sl.no'],
                'description': ['description', 'particulars', 'item', 'work', 'details', 'item of work'],
                'unit': ['unit', 'units', 'measurement', 'measure'],
                'quantity': ['quantity executed', 'quantity', 'qty executed', 'qty', 'executed qty'],
                'rate': ['rate', 'unit rate', 'price', 'cost per unit'],
                'amount': ['amount', 'total amount', 'value', 'total cost'],
                'prev_quantity': ['previous quantity', 'prev qty', 'cumulative qty', 'upto date qty'],
                'remark': ['remark', 'remarks', 'note', 'comment']
            }
            
            # Map columns with flexibility
            mapped_columns = {}
            for target_col, possible_names in column_mappings.items():
                for possible_name in possible_names:
                    for actual_col in df.columns:
                        if possible_name in actual_col or actual_col in possible_name:
                            mapped_columns[target_col] = actual_col
                            break
                    if target_col in mapped_columns:
                        break
            
            bill_items = []
            for index, row in df.iterrows():
                # Skip rows with no meaningful data
                description = clean_text(row.get(mapped_columns.get('description', ''), ''))
                quantity = safe_float_conversion(row.get(mapped_columns.get('quantity', ''), 0))
                
                # Only process items with non-zero quantity and valid description
                if not description or quantity <= 0:
                    continue
                
                # Get rate value
                rate = safe_float_conversion(row.get(mapped_columns.get('rate', ''), 0))
                
                # Handle blank/zero rate logic according to VBA specification
                # If rate is blank or zero, only populate serial_no and description
                if rate == 0 or pd.isna(row.get(mapped_columns.get('rate', ''))) or str(row.get(mapped_columns.get('rate', ''))).strip() == '':
                    item = {
                        'serial_no': clean_text(row.get(mapped_columns.get('serial_no', ''), str(index + 1))),
                        'description': description,
                        'unit': '',  # Keep blank when rate is blank/zero
                        'quantity': 0,  # Keep zero when rate is blank/zero
                        'rate': 0,  # Keep zero when rate is blank/zero
                        'amount': 0,  # Keep zero when rate is blank/zero
                        'remark': ''  # Keep blank when rate is blank/zero
                    }
                else:
                    # Normal processing when rate is not blank/zero
                    item = {
                        'serial_no': clean_text(row.get(mapped_columns.get('serial_no', ''), str(index + 1))),
                        'description': description,
                        'unit': clean_text(row.get(mapped_columns.get('unit', ''), ''),
                        'quantity': quantity,
                        'rate': rate,
                        'remark': clean_text(row.get(mapped_columns.get('remark', ''), ''))
                    }
                    
                    # Handle amount calculation
                    if 'amount' in mapped_columns:
                        amount = safe_float_conversion(row.get(mapped_columns['amount'], 0))
                        item['amount'] = amount if amount > 0 else (item['quantity'] * item['rate'])
                    else:
                        item['amount'] = item['quantity'] * item['rate']
                    
                    # Apply specific rounding rules for bill quantities
                    item['quantity'] = round_to_nearest(item['quantity'], 2)
                    item['rate'] = round_to_nearest(item['rate'], 2)
                    item['amount'] = round_to_nearest(item['amount'], 2)
                
                bill_items.append(item)
            
            logger.info(f"Processed {len(bill_items)} bill quantity items")
            return bill_items
            
        except Exception as e:
            logger.error(f"Error processing bill quantity sheet: {str(e)}")
            return []

    def process_extra_items_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Process extra items sheet with enhanced validation"""
        try:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            
            # Clean column names
            df.columns = df.columns.str.strip().str.lower()
            
            # Column mapping specific to extra items
            column_mappings = {
                'serial_no': ['s.no', 'serial', 'sr.no', 'no', 'item no'],
                'description': ['description', 'particulars', 'item', 'work', 'extra work'],
                'unit': ['unit', 'units', 'measurement'],
                'quantity': ['quantity', 'qty', 'executed qty'],
                'rate': ['rate', 'unit rate', 'approved rate', 'price'],
                'amount': ['amount', 'total', 'value'],
                'approval_ref': ['approval', 'reference', 'sanction', 'order'],
                'remark': ['remark', 'remarks', 'justification']
            }
            
            # Map columns
            mapped_columns = {}
            for target_col, possible_names in column_mappings.items():
                for possible_name in possible_names:
                    for actual_col in df.columns:
                        if possible_name in actual_col:
                            mapped_columns[target_col] = actual_col
                            break
                    if target_col in mapped_columns:
                        break
            
            extra_items = []
            for index, row in df.iterrows():
                # Skip rows with insufficient data
                description = clean_text(row.get(mapped_columns.get('description', ''), '')
                quantity = safe_float_conversion(row.get(mapped_columns.get('quantity', ''), 0))
                
                if not description or quantity <= 0:
                    continue
                
                # Get rate value
                rate = safe_float_conversion(row.get(mapped_columns.get('rate', ''), 0))
                
                # Handle blank/zero rate logic according to VBA specification
                # If rate is blank or zero, only populate serial_no and description
                if rate == 0 or pd.isna(row.get(mapped_columns.get('rate', ''))) or str(row.get(mapped_columns.get('rate', ''))).strip() == '':
                    item = {
                        'serial_no': clean_text(row.get(mapped_columns.get('serial_no', ''), str(index + 1))),
                        'description': description,
                        'unit': '',  # Keep blank when rate is blank/zero
                        'quantity': 0,  # Keep zero when rate is blank/zero
                        'rate': 0,  # Keep zero when rate is blank/zero
                        'amount': 0,  # Keep zero when rate is blank/zero
                        'approval_ref': '',  # Keep blank when rate is blank/zero
                        'remark': ''  # Keep blank when rate is blank/zero
                    }
                else:
                    # Normal processing when rate is not blank/zero
                    item = {
                        'serial_no': clean_text(row.get(mapped_columns.get('serial_no', ''), str(index + 1))),
                        'description': description,
                        'unit': clean_text(row.get(mapped_columns.get('unit', ''), ''),
                        'quantity': quantity,
                        'rate': rate,
                        'approval_ref': clean_text(row.get(mapped_columns.get('approval_ref', ''), '')),
                        'remark': clean_text(row.get(mapped_columns.get('remark', ''), ''))
                    }
                    
                    # Calculate amount
                    if 'amount' in mapped_columns:
                        amount = safe_float_conversion(row.get(mapped_columns['amount'], 0))
                        item['amount'] = amount if amount > 0 else (item['quantity'] * item['rate'])
                    else:
                        item['amount'] = item['quantity'] * item['rate']
                    
                    # Apply rounding rules
                    item['quantity'] = round_to_nearest(item['quantity'], 2)
                    item['rate'] = round_to_nearest(item['rate'], 2)
                    item['amount'] = round_to_nearest(item['amount'], 2)
                
                extra_items.append(item)
            
            logger.info(f"Processed {len(extra_items)} extra items")
            return extra_items
            
        except Exception as e:
            logger.error(f"Error processing extra items sheet: {str(e)}")
            return []

    def calculate_financial_totals(self, processed_data: Dict[str, Any]) -> Dict[str, float]:
        """Calculate comprehensive financial totals with all required calculations"""
        try:
            totals = {
                'bill_quantity_total': 0.0,
                'extra_items_total': 0.0,
                'grand_total': 0.0,
                'gst_rate': 18.0,  # Standard GST rate
                'gst_amount': 0.0,
                'total_with_gst': 0.0,
                'premium_amount': 0.0,
                'net_payable': 0.0
            }
            
            # Calculate bill quantity total
            bill_items = processed_data.get('bill_quantity', [])
            totals['bill_quantity_total'] = sum(item.get('amount', 0) for item in bill_items)
            
            # Calculate extra items total
            extra_items = processed_data.get('extra_items', [])
            totals['extra_items_total'] = sum(item.get('amount', 0) for item in extra_items)
            
            # Calculate grand total (before taxes)
            totals['grand_total'] = totals['bill_quantity_total'] + totals['extra_items_total']
            
            # Handle premium calculation if present in work order data
            work_order_data = processed_data.get('work_order', [])
            if isinstance(work_order_data, dict) and 'tender_premium' in work_order_data:
                premium_percent = safe_float_conversion(work_order_data['tender_premium'])
                totals['premium_amount'] = totals['grand_total'] * (premium_percent / 100)
                totals['grand_total'] += totals['premium_amount']
            
            # Calculate GST
            totals['gst_amount'] = totals['grand_total'] * (totals['gst_rate'] / 100)
            totals['total_with_gst'] = totals['grand_total'] + totals['gst_amount']
            
            # Net payable (final amount)
            totals['net_payable'] = totals['total_with_gst']
            
            # Apply rounding rules
            for key in totals:
                if isinstance(totals[key], float):
                    totals[key] = round_to_nearest(totals[key], 2)
            
            logger.info(f"Calculated financial totals: Grand total = ₹{totals['grand_total']:,.2f}")
            return totals
            
        except Exception as e:
            logger.error(f"Error calculating financial totals: {str(e)}")
            return {}

    def generate_note_sheet_content(self, processed_data: Dict[str, Any]) -> str:
        """
        Generate Note Sheet content based on VBA logic in finally approve bill notes.txt
        This method creates automated notes similar to the VBA implementation
        """
        try:
            # Extract required data from processed data
            title_data = processed_data.get('title', {})
            totals = processed_data.get('totals', {})
            
            work_order_amount = totals.get('work_order_total', 0) or totals.get('grand_total', 0)
            bill_amount = totals.get('bill_quantity_total', 0) or totals.get('grand_total', 0)
            
            # Calculate percentage of work done
            percentage_work_done = (bill_amount / work_order_amount * 100) if work_order_amount > 0 else 0
            
            # Initialize notes
            notes = []
            serial_number = 1
            
            # Note 1: Work completion percentage
            notes.append(f"{serial_number}. The work has been completed {percentage_work_done:.2f}% of the Work Order Amount.")
            serial_number += 1
            
            # Note 2: Conditional notes based on work percentage
            if percentage_work_done < 90:
                notes.append(f"{serial_number}. The execution of work at final stage is less than 90% of the Work Order Amount, the Requisite Deviation Statement is enclosed to observe check on unuseful expenditure. Approval of the Deviation is having jurisdiction under this office.")
                serial_number += 1
            elif percentage_work_done > 100 and percentage_work_done <= 105:
                notes.append(f"{serial_number}. Requisite Deviation Statement is enclosed. The Overall Excess is less than or equal to 5% and is having approval jurisdiction under this office.")
                serial_number += 1
            elif percentage_work_done > 105:
                notes.append(f"{serial_number}. Requisite Deviation Statement is enclosed. The Overall Excess is more than 5% and Approval of the Deviation Case is required from the Superintending Engineer, PWD Electrical Circle, Udaipur.")
                serial_number += 1
            
            # Note 3: Extra items conditions (if extra items exist)
            extra_items = processed_data.get('extra_items', [])
            if extra_items:
                extra_items_total = totals.get('extra_items_total', 0)
                if work_order_amount > 0:
                    extra_item_percentage = (extra_items_total / work_order_amount) * 100
                else:
                    extra_item_percentage = 0
                
                if extra_item_percentage > 5:
                    notes.append(f"{serial_number}. The amount of Extra items is Rs. {extra_items_total:,.2f}. which is {extra_item_percentage:.2f} % of the Work Order Amount; exceed 5%, require approval from the Superintending Engineer, PWD Electrical Circle, Udaipur.")
                else:
                    notes.append(f"{serial_number}. The amount of Extra items is Rs. {extra_items_total:,.2f}. which is {extra_item_percentage:.2f} % of the Work Order Amount; under 5%, approval of the same is to be granted by this office.")
                serial_number += 1
            
            # Note 4: Mandatory conditions
            notes.append(f"{serial_number}. Quality Control (QC) test reports attached.")
            serial_number += 1
            
            notes.append(f"{serial_number}. A copy of Hand Over statement duly signed by client department representative is attached.")
            serial_number += 1
            
            notes.append(f"{serial_number}. Please peruse above details for necessary decision-making.")
            serial_number += 1
            
            # Signature block
            notes.append("")
            notes.append("                                Premlata Jain")
            notes.append("                               AAO- As Auditor")
            
            # Join all notes with newlines
            note_content = "\n".join(notes)
            
            logger.info("Generated Note Sheet content successfully")
            return note_content
            
        except Exception as e:
            logger.error(f"Error generating Note Sheet content: {str(e)}")
            # Return a default note if generation fails
            return "Error generating notes. Please check the data and try again."
    
    def get_processing_summary(self) -> Dict[str, Any]:
        """Get summary of processing results for reporting"""
        return {
            'file_pattern': self.file_pattern,
            'sheets_processed': len(self.processed_data),
            'validation_warnings': self.validation_warnings,
            'total_items': sum(len(items) if isinstance(items, list) else 0 
                              for items in self.processed_data.values()),
            'has_financial_data': 'totals' in self.processed_data,
            'processing_timestamp': datetime.now().isoformat()
        }
